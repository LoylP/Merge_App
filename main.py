import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

def extract_bold_services(file, name_service):
    wb = load_workbook(file)
    ws = wb.active

    start_row = None
    for row in ws.iter_rows():
        if row[0].value and "Dịch vụ-Loại dịch vụ" in str(row[0].value):
            start_row = row[0].row
            break

    bold_rows = []
    if start_row:
        for row in ws.iter_rows(min_row=start_row + 1, values_only=False):  
            cell = row[0]  
            if cell.font and cell.font.bold:  
                row_data = [c.value for c in row]  
                row_data.append(name_service)  
                bold_rows.append(row_data)
    
    return bold_rows

def extract_department_data(file):
    wb = load_workbook(file)
    ws = wb.active
    
    start_row = None
    for row in ws.iter_rows():
        if row[0].value and "Theo phòng ban" in str(row[0].value):
            start_row = row[0].row
            break
    
    department_data = []
    if start_row:
        for row in ws.iter_rows(min_row=start_row + 1, values_only=True):  
            if row[0] is None:
                break
            
            dept_name = row[0]
            total_requests = row[1]
            sla_requests = row[2]
            failed_requests = total_requests - sla_requests if total_requests is not None and sla_requests is not None else None
            sla_percentage = (sla_requests / total_requests * 100) if total_requests and sla_requests is not None else None
            
            department_data.append([dept_name, total_requests, sla_requests, failed_requests, sla_percentage])  
    
    return department_data

st.title("Merge Excel File") 

uploaded_files = st.file_uploader("Tải lên các file Excel", accept_multiple_files=True, type=["xlsx"])

file_names = []
if uploaded_files:
    for i, file in enumerate(uploaded_files):
        name = st.text_input(f"Nhập tên cho file {i+1}", f"File_{i+1}")
        file_names.append((file, name))

if st.button("Xử lý và Xuất File") and uploaded_files:
    bold_rows_all = []
    department_data = None
    
    for i, (file, name) in enumerate(file_names):
        if i == 0:
            department_data = extract_department_data(file)
        bold_rows_all.extend(extract_bold_services(file, name))
    
    df_services = pd.DataFrame(bold_rows_all)
    df_services.columns = ["Dịch vụ-Loại dịch vụ", "Tổng SL yêu cầu", "SL yêu cầu đạt SLA", "SL yêu cầu không đạt SLA", "Tỉ lệ", "Column12"]
    df_services = df_services.drop_duplicates().reset_index(drop=True)
    
    df_departments = pd.DataFrame(department_data, columns=["Theo phòng ban", "Tổng SL Yêu cầu", "SL Yêu cầu đạt SLA", "SL Yêu cầu không đạt SLA", "Tỉ lệ"])
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_departments.to_excel(writer, index=False, startrow=0, startcol=0)
        start_row = len(df_departments) + 3
        df_services.to_excel(writer, index=False, startrow=start_row, startcol=0)
    
    output.seek(0)
    st.success("✅ Xuất file thành công!")
    st.download_button(label="Tải file xuống", data=output, file_name="BaoCaoSLA.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
  


  